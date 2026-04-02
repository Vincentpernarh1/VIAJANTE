import pandas as pd
import openpyxl

print("Exporting VIAJANTE CADASTRO_MDR to Excel for comparison...")

wb = openpyxl.load_workbook('BD/VIAJANTE DHL.xlsm', data_only=True, keep_vba=False)
ws = wb['CADASTRO_MDR']

# Read all data
headers = [cell.value for cell in ws[1]]
data = []
for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
    if row[4]:  # Has MDR value
        data.append(row)

df = pd.DataFrame(data, columns=headers)

# Export to Excel
output_file = 'VIAJANTE_CADASTRO_MDR_EXPORT.xlsx'
df.to_excel(output_file, index=False, sheet_name='CADASTRO_MDR')

print(f"✅ Exported to {output_file}")
print(f"   Total rows: {len(df)}")
print(f"   Unique MDRs: {df['MDR'].nunique()}")
print(f"\n💡 Now compare this file with BD/BD_CADASTRO_MDR.xlsx")
print(f"   Look for:")
print(f"   1. MDR 4315 rows (should exist for your supplier)")
print(f"   2. Different capacity values for same MDR")
print(f"   3. EMPILHAMENTO MÁXIMO values")
