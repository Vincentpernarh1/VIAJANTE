import math

import pandas as pd
from tkinter import *
from tkinter import ttk
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from math import ceil
import re
from PIL import Image, ImageTk
import traceback
import os
import glob
import numpy as np
import warnings 
import contextlib
import unicodedata

# Suppress xlrd / Excel warnings
warnings.simplefilter("ignore")

# Optionally suppress all print output from the engine
with contextlib.redirect_stdout(None), contextlib.redirect_stderr(None):
    df = pd.read_excel("Template.xlsx", engine="openpyxl")
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=".*OLE2.*"
)
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=".*CODEPAGE.*"
)
warnings.filterwarnings(
    "ignore",
    message="^WARNING .*" # Hides the file size warnings which don't have a category
)



caminho_base = os.getcwd()

# Helper function to find latest file matching pattern
def get_latest_file(pattern, fallback=None):
    """Find the most recent file matching the pattern, or return fallback file if none found.
    
    Args:
        pattern: Glob pattern to match files (e.g., "BD/BD_CADASTRO_PN_*.xlsx")
        fallback: Fallback filename if no dated file is found (e.g., "BD/BD_CADASTRO_PN.xlsx")
        
    Returns:
        str: Path to the latest file, or fallback, or None
    """
    files = glob.glob(pattern)
    if files:
        # Sort by modification time, newest first
        files.sort(key=os.path.getmtime, reverse=True)
        return files[0]
    elif fallback and os.path.exists(fallback):
        return fallback
    return None

# Lista global para coletar erros e avisos para mostrar ao usuário
erros_processamento = []

def adicionar_erro(mensagem, tipo="ERRO"):
    """Adiciona uma mensagem de erro ou aviso à lista global sem duplicatas."""
    msg = f"[{tipo}] {mensagem}"
    if msg not in erros_processamento:
        erros_processamento.append(msg)
        print(msg)

def limpar_erros():
    """Limpa a lista de erros"""
    global erros_processamento
    erros_processamento = []

def obter_erros():
    """Retorna a lista de erros acumulados"""
    return erros_processamento.copy()


def normalize_sheet_name(desired_name, available_sheets):
    """
    Finds a matching sheet name from available_sheets that matches desired_name,
    ignoring case and accents.
    
    Args:
        desired_name: The sheet name we want (e.g., 'Sábado', 'Geral', 'Domingo')
        available_sheets: List of actual sheet names in the Excel file
    
    Returns:
        The actual sheet name from the file that matches, or None if no match found
    """
    if desired_name is None or not available_sheets:
        return None
    
    def remove_accents(text):
        """Remove accents from text"""
        if text is None:
            return ''
        text = str(text)
        # Normalize to NFD (decomposed form) and filter out combining characters
        nfd = unicodedata.normalize('NFD', text)
        return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')
    
    # Normalize the desired name (remove accents, convert to upper)
    desired_normalized = remove_accents(desired_name).upper().strip()
    
    # Try to find a match
    for sheet in available_sheets:
        sheet_normalized = remove_accents(str(sheet)).upper().strip()
        if sheet_normalized == desired_normalized:
            return sheet  # Return the actual sheet name from the file
    
    return None  # No match found
    

def Processar_Demandas(cod_destino, pasta_demandas="Demandas", sheet_name=None):
    
    # Define o caminho completo para a pasta de demandas
    caminho_pasta = os.path.join(caminho_base, pasta_demandas)

    # Verifica se a pasta de demandas existe
    if not os.path.isdir(caminho_pasta):
        adicionar_erro(f"Pasta de demandas não encontrada: '{caminho_pasta}'", "ERRO")
        return pd.DataFrame()

    # Lista para armazenar os DataFrames de cada arquivo processado
    lista_dfs = []

    # Percorre todos os arquivos na pasta de demandas
    for nome_arquivo in os.listdir(caminho_pasta):
        caminho_completo_arquivo = os.path.join(caminho_pasta, nome_arquivo)
        nome_arquivo_lower = nome_arquivo.lower()
        
        try:
            # --- MANTÉM A LÓGICA ORIGINAL PARA ARQUIVOS .TXT E .CSV ---
            if nome_arquivo_lower.endswith((".txt", ".csv")):
                dados_arquivo_atual = []
                with open(caminho_completo_arquivo, "r", encoding="utf-8", errors="ignore") as arquivo:
                    linhas_a_processar = arquivo.readlines()

                # Processa cada linha extraída do arquivo de texto
                for linha in linhas_a_processar:
                    if "AUTOMATIC" in linha:
                        continue

                    linha = linha.strip()

                    # A lógica de fatiamento requer um comprimento mínimo
                    if len(linha) >= 20:
                        try:
                            # Extrai os dados com base na posição dos caracteres
                            desenho = linha[3:14]
                            cod_fornecedor = linha[-20:-11]
                            quantidade = linha[-11:].replace("+", "")

                            # Adiciona os dados extraídos à lista deste arquivo
                            dados_arquivo_atual.append({
                                "DESENHO": int(desenho.strip()),
                                "COD FORNECEDOR": int(cod_fornecedor.strip()),
                                "QTDE": int(quantidade.strip()),
                            })
                        except (ValueError, IndexError):
                            # Ignora linhas que não seguem o formato esperado
                            continue
                
                # Se dados foram extraídos do arquivo, cria um DataFrame
                if dados_arquivo_atual:
                    df_temp = pd.DataFrame(dados_arquivo_atual)
                    if cod_destino is not None:
                        df_temp["COD DESTINO"] = cod_destino
                    df_temp['IS_FLECHINHA'] = 0  # .txt files are NOT flechinha
                    lista_dfs.append(df_temp)

            # --- NOVA LÓGICA PARA PROCESSAR ARQUIVOS EXCEL (.XLS, .XLSX) ---
            elif nome_arquivo_lower.endswith((".xls", ".xlsx")) and ("saturação" not in nome_arquivo_lower and "saturacao" not in nome_arquivo_lower) and not nome_arquivo_lower.startswith("~$"):
                               
                # Mapeamento dos nomes de coluna do arquivo Excel para os nomes desejados
                colunas_mapeamento = {
                    'DESENHO': 'DESENHO',
                    'COD ORIGEM': 'COD FORNECEDOR',
                    'ENTREGA SOLICITADA': 'QTDE',
                    'COD DESTINO': 'COD DESTINO'
                }
                
                # Lê o arquivo Excel
                df_excel = pd.read_excel(caminho_completo_arquivo)

                # Pega a lista de colunas que precisamos do arquivo original
                colunas_originais_necessarias = list(colunas_mapeamento.keys())

                # Verifica se todas as colunas necessárias existem no arquivo
                if not all(coluna in df_excel.columns for coluna in colunas_originais_necessarias):
                    faltando = [c for c in colunas_originais_necessarias if c not in df_excel.columns]
                    adicionar_erro(f"Arquivo '{nome_arquivo}': Colunas faltando: {', '.join(faltando)}", "AVISO")
                    continue

                # 1. Seleciona apenas as colunas que nos interessam
                df_temp = df_excel[colunas_originais_necessarias].copy()
                
                # 2. Renomeia as colunas para o padrão final
                df_temp.rename(columns=colunas_mapeamento, inplace=True)

                # 3. Filtra por COD DESTINO se fornecido
                if cod_destino is not None:
                    df_temp = df_temp[df_temp['COD DESTINO'].astype(str) == str(cod_destino)]

                # 4. Marca como NÃO FLECHINHA (Excel normal, não saturação)
                df_temp['IS_FLECHINHA'] = 0
                
                # 5. Adiciona o DataFrame processado à lista para concatenação posterior
                lista_dfs.append(df_temp)
                
            elif nome_arquivo_lower.endswith((".xls", ".xlsx")) and ("saturação" in nome_arquivo_lower or "saturacao" in nome_arquivo_lower)  and not nome_arquivo_lower.startswith("~$"):
                                
                # Só processa arquivos de saturação se sheet_name foi fornecido
                if sheet_name is None:
                    continue
                
                try:
                    xl_file = pd.ExcelFile(caminho_completo_arquivo)
                    
                    # Normaliza o nome da sheet para encontrar correspondência
                    actual_sheet_name = normalize_sheet_name(sheet_name, xl_file.sheet_names)
                    
                    # Verifica se a sheet existe
                    if actual_sheet_name is None:
                        adicionar_erro(f"Arquivo '{nome_arquivo}': Sheet '{sheet_name}' não encontrada. Disponíveis: {', '.join(xl_file.sheet_names)}", "ERRO")
                        continue
                    
                except Exception as e:
                    continue
                
                # Lê o arquivo Excel de saturação da sheet específica com header na linha 3 (índice 2)
                df_excel = pd.read_excel(caminho_completo_arquivo, sheet_name=actual_sheet_name, header=2)
                
                colunas_saturacao_mapeamento = {}
                
                # Procura pelas colunas necessárias
                for col in df_excel.columns:
                    col_upper = str(col).upper().strip()
                    if 'DESENHO' in col_upper and 'FIAT' in col_upper:
                        colunas_saturacao_mapeamento[col] = 'DESENHO'
                    elif 'CÓDIGO' in col_upper and 'IMS' in col_upper:
                        # IMS será mantido como COD IMS
                        colunas_saturacao_mapeamento[col] = 'COD IMS'
                    elif 'QUANTIDADE' in col_upper and 'SOLICITADA' in col_upper:
                        colunas_saturacao_mapeamento[col] = 'QTDE'
                
                # Verifica se encontrou as colunas OBRIGATÓRIAS (DESENHO e QTDE)
                tem_desenho = 'DESENHO' in colunas_saturacao_mapeamento.values()
                tem_qtde = 'QTDE' in colunas_saturacao_mapeamento.values()
                
                if not (tem_desenho and tem_qtde):
                    faltando = []
                    if not tem_desenho: faltando.append('DESENHO FIAT')
                    if not tem_qtde: faltando.append('QUANTIDADE SOLICITADA')
                    adicionar_erro(f"Arquivo saturação '{nome_arquivo}': Colunas obrigatórias faltando: {', '.join(faltando)}", "ERRO")
                    continue

                # 1. Seleciona apenas as colunas que nos interessam
                colunas_originais_necessarias = list(colunas_saturacao_mapeamento.keys())
                df_temp = df_excel[colunas_originais_necessarias].copy()
                
                # 2. Renomeia as colunas para o padrão final
                df_temp.rename(columns=colunas_saturacao_mapeamento, inplace=True)
                
                
                
                # 3. Remove .0 do final dos valores de COD IMS (converte para int, depois para string)
                if 'COD IMS' in df_temp.columns:
                    df_temp['COD IMS'] = pd.to_numeric(df_temp['COD IMS'], errors='coerce')
                    df_temp['COD IMS'] = df_temp['COD IMS'].fillna(0).astype(int).astype(str)
                    # df_temp['COD IMS'] = df_temp['COD IMS'].replace('0', pd.NA)  # Restaura NaN onde era 0
                    
                
                # 4. COD FORNECEDOR sempre nulo para arquivos de saturação (será preenchido depois via COD IMS)
                df_temp['COD FORNECEDOR'] = ""
                # print(f"INFO: Coluna 'COD FORNECEDOR' definida como nula (será derivada de COD IMS).")

                # 5. Adiciona a coluna COD DESTINO
                if cod_destino is not None:
                    df_temp["COD DESTINO"] = cod_destino
                
                # 6. Marca como FLECHINHA (saturação file)
                df_temp['IS_FLECHINHA'] = 1
                
                # 7. Adiciona o DataFrame processado à lista para concatenação posterior
                lista_dfs.append(df_temp)

        except Exception as e:
            adicionar_erro(f"Erro ao processar arquivo '{nome_arquivo}': {str(e)}", "ERRO")
            continue

    # --- LÓGICA FINAL PARA CONSOLIDAR OS DADOS ---
    # Se a lista de DataFrames estiver vazia, retorna um DataFrame vazio
    if not lista_dfs:
        adicionar_erro("Nenhum dado válido foi processado. Verifique os arquivos na pasta Demandas.", "ERRO")
        return pd.DataFrame()
    
    # Concatena todos os DataFrames da lista em um único DataFrame final
    df_final = pd.concat(lista_dfs, ignore_index=True)
    
    # Ensure COD IMS column always exists (for files without it, fill with NaN)
    if 'COD IMS' not in df_final.columns:
        df_final['COD IMS'] = pd.NA

    # Para arquivos de saturação, COD FORNECEDOR pode estar nulo
    # Neste caso, usamos COD IMS como COD FORNECEDOR
    if 'COD IMS' in df_final.columns:
        # Preenche COD FORNECEDOR com COD IMS onde estiver nulo
        df_final['COD FORNECEDOR'] = df_final['COD FORNECEDOR'].fillna(df_final['COD IMS'])

    # Garante que as colunas numéricas tenham o tipo de dados correto, tratando possíveis erros
    # COD FORNECEDOR is intentionally excluded here because fillna(COD IMS) may have placed
    # compound codes like "56589/46051" into it — converting those to numeric would produce NaN.
    # COD FORNECEDOR is cleaned separately below to preserve compound values.
    colunas_numericas = ["DESENHO", "QTDE"]
    for col in colunas_numericas:
        if col in df_final.columns:
            df_final[col] = pd.to_numeric(df_final[col], errors='coerce')

    # Remove linhas onde a conversão numérica falhou (resultando em NaT/NaN)
    # Mas mantém se apenas COD FORNECEDOR estiver nulo (pode vir de COD IMS depois)
    df_final.dropna(subset=["DESENHO", "QTDE"], inplace=True)

    # Remove linhas onde QTDE é zero ou negativo
    df_final = df_final[df_final['QTDE'] > 0]

    # Converte DESENHO e QTDE para inteiro após remover os nulos
    for col in colunas_numericas:
        if col in df_final.columns:
            df_final[col] = pd.to_numeric(df_final[col], errors='coerce')
            # Só converte para int se não tiver NaN
            if df_final[col].notna().all():
                df_final[col] = df_final[col].astype(int)

    # Clean up COD FORNECEDOR: always store as string to prevent float upcast.
    # Numeric values drop the .0 suffix; compound codes kept; empty -> None (not '0',
    # because '0' would cause false substring matches in main.py matching logic).
    if 'COD FORNECEDOR' in df_final.columns:
        def _clean_forn(val):
            s = str(val).strip()
            if s in ('nan', '', 'None'):
                return None
            if '/' in s:
                return s
            try:
                return str(int(float(s)))
            except (ValueError, TypeError):
                return s
        df_final['COD FORNECEDOR'] = df_final['COD FORNECEDOR'].apply(_clean_forn)
    
    # Clean up COD DESTINO to remove .0 suffix if it exists
    if 'COD DESTINO' in df_final.columns:
        df_final['COD DESTINO'] = df_final['COD DESTINO'].astype(str).str.replace(r'\.0$', '', regex=True)
    
    # Clean up COD IMS to remove .0 suffix if it exists  
    if 'COD IMS' in df_final.columns:
        df_final['COD IMS'] = df_final['COD IMS'].astype(str).str.replace(r'\.0$', '', regex=True)
    
    return df_final

# Exemplo de como chamar a função
# df_processado = Processar_Demandas(cod_destino="BR01")
# print(df_processado)







def desenhar_caminhoes(canvas, ocupacao, caminhao_img):
    canvas.delete("all")

    if caminhao_img is None:
        return

    quad_por_caminhao = 35
    quad_linha = 7
    quad_coluna = 5
    quad_largura = 14
    quad_altura = 14
    margem = 10

    total_quads = ceil(ocupacao * quad_por_caminhao / 100)
    max_caminhoes = 3
    num_caminhoes = min((total_quads - 1) // quad_por_caminhao + 1, max_caminhoes)

    for caminhao_idx in range(num_caminhoes):
        # Posição em "grade" 2 acima, 1 abaixo
        if caminhao_idx < 2:
            x_offset = margem + caminhao_idx * 180  # lado a lado
            y_offset = margem
        else:
            x_offset = margem + 90  # centraliza abaixo dos dois
            y_offset = margem + 130

        canvas.create_image(x_offset + 12, y_offset + 17, image=caminhao_img, anchor=NW)

        x_inicial_grade = x_offset + 50
        y_inicial_grade = y_offset + 10

        for i in range(quad_coluna):
            for j in range(quad_linha):
                idx = caminhao_idx * quad_por_caminhao + (quad_coluna - 1 - i) * quad_linha + j
                x1 = x_inicial_grade + j * quad_largura
                y1 = y_inicial_grade + i * quad_altura
                x2 = x1 + quad_largura
                y2 = y1 + quad_altura
                cor = "#0070C0" if idx < total_quads else "#D9D9D9"
                canvas.create_rectangle(x1, y1, x2, y2, fill=cor, outline='black')



def calcular_empilhamento_line_haul(df_saturacao, db_empilhamento):
    empilhamento_rows = []

    base_df = df_saturacao[df_saturacao['EMBALAGEM_BASE'] == 1]
    sobre_df = df_saturacao[df_saturacao['EMBALAGEM_SOBREPOSTA'] == 1]

    for _, base_row in base_df.iterrows():
        for _, sobre_row in sobre_df.iterrows():
            if base_row['COD FORNECEDOR'] == sobre_row['COD FORNECEDOR']:
                fornecedor = base_row['COD FORNECEDOR']
                embal_base = base_row['EMBALAGEM']
                embal_sobre = sobre_row['EMBALAGEM']

                empilhamento_match = db_empilhamento[
                    (db_empilhamento['COD FORNECEDOR'] == fornecedor) &
                    (db_empilhamento['MDR BASE'] == embal_base) &
                    (db_empilhamento['MDR SOBREPOSTA'] == embal_sobre)
                ]

                if empilhamento_match.empty:
                    continue

                capacidade_veiculo = base_row['CAPACIDADE']

                total_base = base_row['TOTAL DE CXS']
                total_sobre = sobre_row['TOTAL DE CXS']

                usadas_base = 0
                usadas_sobre = 0

                # Empilha 1 base com 1 sobreposta (não considera EMPILHAMENTO BASE)
                while total_base >= 1 and total_sobre >= 1:
                    total_base -= 1
                    total_sobre -= 1
                    usadas_base += 1
                    usadas_sobre += 1

                total_empilhado = usadas_base + usadas_sobre
                chave = f"{fornecedor}-{embal_base}-{embal_sobre}"
               
                saturacao = total_empilhado / capacidade_veiculo

                empilhamento_rows.append({
                    'FORNECEDOR': fornecedor,
                    'EMBALAGEM_BASE': embal_base,
                    'EMBALAGEM_SOBREPOSTA': embal_sobre,
                    'CAPACIDADE_VEÍCULO': capacidade_veiculo,
                    'TOTAL_DE_EMBALAGENS_BASE': base_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA': sobre_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_BASE_PARA_COMBINAR': usadas_base,
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA_PARA_COMBINAR': usadas_sobre,
                    'EMBALAGENS_BASE_RESTANTE': total_base,
                    'EMBALAGENS_SOBREPOSTA_RESTANTE': total_sobre,
                    'CHAVE': chave,
                    'TOTAL_EMBALAGENS_EMPILHADAS': total_empilhado,
                    'SATURAÇÃO': saturacao,
                    'EMPILHAMENTO BASE': 1  # fixo, pois é 1:1
                })

    return pd.DataFrame(empilhamento_rows)


def calcular_empilhamento(df_saturacao, db_empilhamento):
    empilhamento_rows = []

    base_df = df_saturacao[df_saturacao['EMBALAGEM_BASE'] == 1]
    sobre_df = df_saturacao[df_saturacao['EMBALAGEM_SOBREPOSTA'] == 1]

    for _, base_row in base_df.iterrows():
        for _, sobre_row in sobre_df.iterrows():
            if base_row['COD FORNECEDOR'] == sobre_row['COD FORNECEDOR']:
                fornecedor = base_row['COD FORNECEDOR']
                embal_base = base_row['EMBALAGEM']
                embal_sobre = sobre_row['EMBALAGEM']

                empilhamento_match = db_empilhamento[
                    (db_empilhamento['COD FORNECEDOR'] == fornecedor) &
                    (db_empilhamento['MDR BASE'] == embal_base) &
                    (db_empilhamento['MDR SOBREPOSTA'] == embal_sobre)
                ]

                if empilhamento_match.empty:
                    continue

                emp_base = empilhamento_match.iloc[0]['EMPILHAMENTO BASE']
                capacidade_veiculo = base_row['CAPACIDADE']

                total_base = base_row['TOTAL DE CXS']
                total_sobre = sobre_row['TOTAL DE CXS']

                usadas_base = 0
                usadas_sobre = 0

                while total_base >= emp_base and total_sobre >= 1:
                    total_base -= emp_base
                    total_sobre -= 1
                    usadas_base += emp_base
                    usadas_sobre += 1

                total_empilhado = usadas_base + usadas_sobre
                chave = f"{fornecedor}-{embal_base}-{embal_sobre}"
                
                saturacao = total_empilhado / capacidade_veiculo

                empilhamento_rows.append({
                    'FORNECEDOR': fornecedor,
                    'EMBALAGEM_BASE': embal_base,
                    'EMBALAGEM_SOBREPOSTA': embal_sobre,
                    'CAPACIDADE_VEÍCULO': capacidade_veiculo,
                    'TOTAL_DE_EMBALAGENS_BASE': base_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA': sobre_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_BASE_PARA_COMBINAR': usadas_base,
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA_PARA_COMBINAR': usadas_sobre,
                    'EMBALAGENS_BASE_RESTANTE': total_base,
                    'EMBALAGENS_SOBREPOSTA_RESTANTE': total_sobre,
                    'CHAVE': chave,
                    'TOTAL_EMBALAGENS_EMPILHADAS': total_empilhado,
                    'SATURAÇÃO': saturacao,
                    'EMPILHAMENTO BASE': emp_base
                })

    return pd.DataFrame(empilhamento_rows)




















def completar_informacoes(tree, veiculo, tree_resumo, canvas_caminhoes, caminhao_img, usar_manual=False,caminho_BD = 'BD'):


    def split_key_logic(code):
        """
        Splits a code by '/'. 
        Returns the second element if a split occurs, otherwise returns the original code.
        """
        # Convert to string just in case, then split
        
        parts = str(code).split('/')
        
        if len(parts) > 1:
            # If the split created more than one part, take the second one (index 1)
            return parts[1].strip() 
            
        else:
            # Otherwise, take the original part (index 0)
            return parts[0].strip()
    try:


        # --- Leitura dos arquivos ---
        template = pd.read_excel('Template.xlsx', dtype={'DESENHO': str})
        template = template[template['QTDE'] > 0]
        
        # Ensure COD IMS column exists (for backward compatibility with files that don't have it)
        if 'COD IMS' not in template.columns:
            template['COD IMS'] = ""
        
        # Clean up COD FORNECEDOR: always store as string to prevent float upcast ("800006330.0").
        # Numeric values drop the .0 suffix; compound codes kept; empty → "0".
        if 'COD FORNECEDOR' in template.columns:
            def _clean_cod_forn_template(val):
                s = str(val).strip()
                if s in ('nan', '', 'None'):
                    return '0'
                if '/' in s:
                    return s
                try:
                    return str(int(float(s)))
                except (ValueError, TypeError):
                    return s
            template['COD FORNECEDOR'] = template['COD FORNECEDOR'].apply(_clean_cod_forn_template)
        
        # Clean up COD DESTINO to remove .0 suffix
        if 'COD DESTINO' in template.columns:
            template['COD DESTINO'] = template['COD DESTINO'].astype(str).str.replace(r'\.0$', '', regex=True)
        
        # Use pattern matching to find latest dated files, with fallback to non-dated versions
        BD_PN = get_latest_file(
            os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_PN_*.xlsx"),
            fallback=os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_PN.xlsx")
        )
        BD_MDR = get_latest_file(
            os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_MDR_*.xlsx"),
            fallback=os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_MDR.xlsx")
        )
        
        if BD_PN is None:
            raise FileNotFoundError("BD_CADASTRO_PN file not found. Please ensure database files are available.")
        if BD_MDR is None:
            raise FileNotFoundError("BD_CADASTRO_MDR file not found. Please ensure database files are available.")
        
        VEÍCULOS = os.path.join(caminho_base,caminho_BD,"VEÍCULOS.xlsx")
        db_empilhamento = os.path.join(caminho_base,caminho_BD,"BD_EMPILHAMENTO_EMBALAGENS.xlsx")
        db_efi = os.path.join(caminho_base,caminho_BD,"BD_CADASTRO_MDR_PERDA_COMPRIMENTO.xlsx")
        PN_CT_path = os.path.join(caminho_base,caminho_BD,"PN_Conta_trabalho.xlsx")
       
        # ------------------Working in the DB structrue------------------
        db_PN = pd.read_excel(BD_PN, sheet_name='BD', dtype={'CÓD. FORNECEDOR': int, 'DESENHO': str})
        db_PN = db_PN.rename(columns={'CÓD. FORNECEDOR': 'COD FORNECEDOR'})
        
        # Filter for EMPRESA = 1, 10.12 (not separate 10 and 12!)
        # Note: EMPRESA 10.12 is a single float value in the database
        if 'EMPRESA' in db_PN.columns:
            db_PN = db_PN[db_PN['EMPRESA'].isin([1, 1.0, 10.12])]
        else:
            print("[WARNING] Column 'EMPRESA' not found in BD_CADASTRO_PN")

        db_MDR = pd.read_excel(BD_MDR, sheet_name='BD')
        db_MDR = db_MDR.rename(columns={'DESCRIÇÃO2': 'DESCRIÇÃO'})
        
        # Filter for EMPRESA = 1, 10.12 (not separate 10 and 12!)
        if 'EMPRESA' in db_MDR.columns:
            db_MDR = db_MDR[db_MDR['EMPRESA'].isin([1, 1.0, 10.12])]
        else:
            print("[WARNING] Column 'EMPRESA' not found in BD_CADASTRO_MDR")

        db_veiculos = pd.read_excel(VEÍCULOS, sheet_name='VEÍCULOS')

        db_empilhamento = pd.read_excel(db_empilhamento, sheet_name='BD')
        db_empilhamento = db_empilhamento.rename(columns={'CÓD. FORNECEDOR': 'COD FORNECEDOR'})

        db_efi = pd.read_excel(db_efi,sheet_name='BD')
        
        # --- Load PN_Conta_trabalho for CT validation ---
        pn_ct_lookup = set()  # Will store (FORNECEDOR, DESENHO) pairs
        try:
            if os.path.exists(PN_CT_path):
                db_pn_ct = pd.read_excel(PN_CT_path)
                # Normalize column names to find Fornecedor, Desenho, Destino
                col_map = {}
                for col in db_pn_ct.columns:
                    col_upper = str(col).upper().strip()
                    if 'FORNECEDOR' in col_upper:
                        col_map['FORNECEDOR'] = col
                    elif 'DESENHO' in col_upper:
                        col_map['DESENHO'] = col
                    elif 'DESTINO' in col_upper:
                        col_map['DESTINO'] = col
                
                # Need FORNECEDOR and DESENHO columns for CT matching
                if all(k in col_map for k in ['FORNECEDOR', 'DESENHO']):
                    for _, row in db_pn_ct.iterrows():
                        # Normalize to string and remove .0 from floats
                        # CT matching uses FORNECEDOR (COD IMS) + DESENHO
                        try:
                            forn = str(int(float(row[col_map['FORNECEDOR']]))) if pd.notna(row[col_map['FORNECEDOR']]) else ''
                        except (ValueError, TypeError):
                            forn = str(row[col_map['FORNECEDOR']]).strip()
                        
                        try:
                            desenho = str(int(float(row[col_map['DESENHO']]))) if pd.notna(row[col_map['DESENHO']]) else ''
                        except (ValueError, TypeError):
                            desenho = str(row[col_map['DESENHO']]).strip()
                        
                        if forn and desenho:
                            pn_ct_lookup.add((forn, desenho))  # (FORNECEDOR/COD_IMS, DESENHO)
                    # print(f"[INFO] Loaded {len(pn_ct_lookup)} Fornecedor-Desenho combinations from PN_Conta_trabalho.xlsx")
                else:
                    adicionar_erro("PN_Conta_trabalho.xlsx: Colunas esperadas não encontradas", "AVISO")
        except Exception as e:
            adicionar_erro(f"Erro ao carregar PN_Conta_trabalho.xlsx: {str(e)}", "AVISO")

        # --- Normalização de tipos ---
        db_PN['DESENHO ATUALIZAÇÃO'] = pd.to_datetime(db_PN['DESENHO ATUALIZAÇÃO'], errors='coerce')
        db_MDR['VOLUME'] = pd.to_numeric(db_MDR['VOLUME'], errors='coerce')
        db_MDR['MDR PESO'] = pd.to_numeric(db_MDR['MDR PESO'], errors='coerce')
        db_PN['PESO (Kg) MATERIAL'] = pd.to_numeric(db_PN['PESO (Kg) MATERIAL'], errors='coerce')
        
        
        
        db_PN = db_PN.sort_values('DESENHO ATUALIZAÇÃO', ascending=False)
        
        # Criar chave composta DESENHO+MDR em db_PN
        db_PN['KEY'] = db_PN['DESENHO'].astype(str) + '_' + db_PN['MDR'].astype(str)

        # --- Mapeamentos únicos para .map() seguros ---
        # Filter out nan values and keep most recent entries for db_PN mappings
        mapa_fornecedores = db_PN.drop_duplicates('COD FORNECEDOR').set_index('COD FORNECEDOR')['FORNECEDOR']

        # Mapas baseados na chave composta - already sorted by DESENHO ATUALIZAÇÃO descending
        # This ensures we always get the most recent non-null values
        mapa_pn = db_PN.drop_duplicates('KEY', keep='first').set_index('KEY')['DESCRIÇÃO']
        mapa_mdr = db_PN.drop_duplicates('KEY', keep='first').set_index('KEY')['MDR']
        
        # For QME and PESO, filter out invalid values before mapping
        db_PN_valid_qme = db_PN[db_PN['QME'].notna() & (db_PN['QME'] > 0)]
        mapa_qme = db_PN_valid_qme.drop_duplicates('KEY', keep='first').set_index('KEY')['QME']
        
        db_PN_valid_peso = db_PN[db_PN['PESO (Kg) MATERIAL'].notna()]
        mapa_peso_pn = db_PN_valid_peso.drop_duplicates('KEY', keep='first').set_index('KEY')['PESO (Kg) MATERIAL']

        # Mapas vindos do db_MDR - filter out nan values BEFORE creating mappings
        db_MDR_valid_desc = db_MDR[db_MDR['DESCRIÇÃO'].notna()]
        mapa_descricao_mdr = db_MDR_valid_desc.drop_duplicates('MDR', keep='first').set_index('MDR')['DESCRIÇÃO']
        
        db_MDR_valid_volume = db_MDR[db_MDR['VOLUME'].notna()]
        mapa_volume = db_MDR_valid_volume.drop_duplicates('MDR', keep='first').set_index('MDR')['VOLUME']
        
        db_MDR_valid_peso = db_MDR[db_MDR['MDR PESO'].notna()]
        mapa_peso_mdr = db_MDR_valid_peso.drop_duplicates('MDR', keep='first').set_index('MDR')['MDR PESO']
        
        mapa_peso_max = db_veiculos.set_index('COD VEICULO')['PESO MAXIMO']

        # keep template VEICULO as-is here; Template.xlsx will be
        # updated by input_demanda() when the user forces manual vehicle.

        # --- Enriquecimento do template ---

        # Passo 1: primeiro trazer MDR pelo DESENHO, para podermos montar a KEY
        # Filter out nan MDR values and use most recent (already sorted by DESENHO ATUALIZAÇÃO)
        db_PN_valid_mdr = db_PN[db_PN['MDR'].notna()]
        
        template['MDR'] = template['DESENHO'].map(
            db_PN_valid_mdr.drop_duplicates('DESENHO', keep='first').set_index('DESENHO')['MDR']
        )


        # Passo 2: agora que já temos MDR no template, podemos montar a KEY
        template['KEY'] = template['DESENHO'].astype(str) + '_' + template['MDR'].astype(str)
        

        # Passo 3: enriquecer com os mapas
        template['PESO_MAXIMO'] = template['VEICULO'].map(mapa_peso_max)
        template['MAP_KEY'] = (template['COD IMS'].fillna(template['COD FORNECEDOR']).astype(str).str.split('/').str[0] )

       
        template['MAP_KEY'] = pd.to_numeric(template['MAP_KEY'], errors='coerce')
        template['FORNECEDOR'] = template['MAP_KEY'].map(mapa_fornecedores)
       
        # Clean up FORNECEDOR column - remove .0 suffix if it exists (when mapping fails, it might keep numeric values)
        if 'FORNECEDOR' in template.columns:
            template['FORNECEDOR'] = template['FORNECEDOR'].astype(str).str.replace(r'\.0$', '', regex=True)
            # If FORNECEDOR is 'nan', replace with empty string
            template['FORNECEDOR'] = template['FORNECEDOR'].replace('nan', '')
       
        template = template.drop(columns=['MAP_KEY'])
       
        template['DESCRIÇÃO MATERIAL'] = template['KEY'].map(mapa_pn)
        template['MDR'] = template['KEY'].map(mapa_mdr)  # reforça MDR correto do KEY
        template['DESCRIÇÃO DA EMBALAGEM'] = template['MDR'].map(mapa_descricao_mdr)
        template['QME'] = template['KEY'].map(mapa_qme)

        # Ensure QME is valid (not zero, not NaN) before division
        template['QME'] = template['QME'].fillna(1)  # Replace NaN with 1 to avoid division issues
        template['QME'] = template['QME'].replace(0, 1)  # Replace 0 with 1 to avoid division by zero
        template['QTD EMBALAGENS'] = np.ceil(template['QTDE'] / template['QME'])
        
        # Clean up any infinity values in QTD EMBALAGENS
        template['QTD EMBALAGENS'] = template['QTD EMBALAGENS'].replace([np.inf, -np.inf], np.nan).fillna(0)

        # Use 3 decimal places for M³ to capture small variations (e.g., 0.036 instead of 0.0)
        template['M³'] = round(template['QTD EMBALAGENS'] * template['MDR'].map(mapa_volume), 3)
        template['PESO MAT'] = round(template['QTDE'] * template['KEY'].map(mapa_peso_pn), 1)
        template['PESO MDR'] = round(template['QTD EMBALAGENS'] * template['MDR'].map(mapa_peso_mdr), 1)
        
        # Garante que NaN e infinitos sejam tratados como 0 antes de somar
        template['M³'] = template['M³'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['PESO MAT'] = template['PESO MAT'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['PESO MDR'] = template['PESO MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['PESO TOTAL'] = round(template['PESO MAT'] + template['PESO MDR'], 1)

        # Final cleanup: re-apply string normalisation in case any operation reintroduced floats
        if 'COD FORNECEDOR' in template.columns:
            template['COD FORNECEDOR'] = template['COD FORNECEDOR'].apply(_clean_cod_forn_template)
        
        # COD DESTINO should be string without .0
        if 'COD DESTINO' in template.columns:
            template['COD DESTINO'] = template['COD DESTINO'].astype(str).str.replace(r'\.0$', '', regex=True)

        # Add MOT column if exists in template (from Template.xlsx)
        if 'MOT' not in template.columns:
            template['MOT'] = None
        
        # Add FLECHINHA column if exists in template (from Template.xlsx)
        if 'FLECHINHA' not in template.columns:
            template['FLECHINHA'] = 0
        
        template = template[['COD FORNECEDOR', 'FORNECEDOR', 'COD DESTINO', 'DESENHO', 'QTDE', 'DESCRIÇÃO MATERIAL',
                             'MDR', 'DESCRIÇÃO DA EMBALAGEM', 'QME', 'QTD EMBALAGENS', 'TIPO SATURACAO',
                             'VEICULO', 'MOT', 'FLECHINHA', 'M³', 'PESO MAT', 'PESO MDR', 'PESO TOTAL', 'PESO_MAXIMO']]
        
        # --- CT Validation: Filter rows based on MOT and PN_Conta_trabalho ---
        template['INCLUDE_IN_CALC'] = True  # Default: include all
        
        if 'MOT' in template.columns and len(pn_ct_lookup) > 0:
            ct_excluded_count = 0
            ct_included_count = 0
            ftl_excluded_count = 0
            ftl_included_count = 0
            
            # Group by (COD FORNECEDOR, COD DESTINO, MOT)
            for (cod_forn, dest, mot), group in template.groupby(['COD FORNECEDOR', 'COD DESTINO', 'MOT']):
                mot_upper = str(mot).strip().upper() if pd.notna(mot) else ''
                
                if mot_upper == 'CT':
                    # For CT rows, check if (COD_IMS, DESENHO) exists in pn_ct_lookup
                    # Include only if found in PN_Conta_trabalho
                    for idx, row in group.iterrows():
                        # Get COD IMS for lookup (this is the Fornecedor in PN_Conta_trabalho)
                        cod_ims = row.get('COD IMS', None)
                        try:
                            cod_ims_str = str(int(float(cod_ims))) if pd.notna(cod_ims) else ''
                        except (ValueError, TypeError):
                            cod_ims_str = str(cod_ims).strip() if cod_ims else ''
                        
                        # Skip CT filtering if no COD IMS (include by default)
                        if not cod_ims_str:
                            ct_included_count += 1
                            continue
                        
                        try:
                            desenho_val = row['DESENHO']
                            desenho_str = str(int(float(desenho_val))) if pd.notna(desenho_val) else ''
                        except (ValueError, TypeError):
                            desenho_str = str(desenho_val).strip() if pd.notna(desenho_val) else ''
                        
                        # Handle compound COD IMS (e.g., "24149/36190")
                        # Check if ANY part of the compound IMS matches in the CT lookup
                        cod_ims_parts = [p.strip() for p in cod_ims_str.split('/')]
                        
                        is_in_ct = False
                        for ims_part in cod_ims_parts:
                            key = (ims_part, desenho_str)  # (COD_IMS_PART, DESENHO)
                            if key in pn_ct_lookup:
                                is_in_ct = True
                                break
                        
                        if not is_in_ct:
                            # PN not in CT file, exclude from CT calculation
                            template.at[idx, 'INCLUDE_IN_CALC'] = False
                            ct_excluded_count += 1
                        else:
                            ct_included_count += 1
                
                elif mot_upper in ['FTL', 'LTL']:
                    # For FTL/LTL rows, EXCLUDE if (COD_IMS, DESENHO) exists in pn_ct_lookup
                    # This prevents CT PNs from appearing in FTL for the same COD IMS
                    for idx, row in group.iterrows():
                        # Get COD IMS for lookup
                        cod_ims = row.get('COD IMS', None)
                        try:
                            cod_ims_str = str(int(float(cod_ims))) if pd.notna(cod_ims) else ''
                        except (ValueError, TypeError):
                            cod_ims_str = str(cod_ims).strip() if cod_ims else ''
                        
                        # Skip CT filtering if no COD IMS (include by default)
                        if not cod_ims_str:
                            ftl_included_count += 1
                            continue
                        
                        try:
                            desenho_val = row['DESENHO']
                            desenho_str = str(int(float(desenho_val))) if pd.notna(desenho_val) else ''
                        except (ValueError, TypeError):
                            desenho_str = str(desenho_val).strip() if pd.notna(desenho_val) else ''
                        
                        # Handle compound COD IMS (e.g., "24149/36190")
                        # Check if ANY part of the compound IMS matches in the CT lookup
                        cod_ims_parts = [p.strip() for p in cod_ims_str.split('/')]
                        
                        is_in_ct = False
                        for ims_part in cod_ims_parts:
                            key = (ims_part, desenho_str)  # (COD_IMS_PART, DESENHO)
                            if key in pn_ct_lookup:
                                is_in_ct = True
                                break
                        
                        if is_in_ct:
                            # PN is in CT file for this COD IMS, so exclude from FTL calculation
                            template.at[idx, 'INCLUDE_IN_CALC'] = False
                            ftl_excluded_count += 1
                        else:
                            ftl_included_count += 1
                # Else: MOT is NaN or other value - skip CT filtering (include by default)
            
            if ct_excluded_count > 0:
                adicionar_erro(f"{ct_excluded_count} PN(s) com MOT=CT excluídos (não encontrados em PN_Conta_trabalho)", "INFO")
            if ct_included_count > 0:
                adicionar_erro(f"{ct_included_count} PN(s) com MOT=CT incluídos (encontrados em PN_Conta_trabalho)", "INFO")
            if ftl_excluded_count > 0:
                adicionar_erro(f"{ftl_excluded_count} PN(s) com MOT!=CT excluídos (encontrados em PN_Conta_trabalho, pertencem ao CT)", "INFO")
            if ftl_included_count > 0:
                adicionar_erro(f"{ftl_included_count} PN(s) com MOT!=CT incluídos (não encontrados em PN_Conta_trabalho)", "INFO")

        # --- Remove duplicates BEFORE calculations ---
        # Deduplicate after enrichment but before saturação calculations
        # This ensures all calculations (volume, peso, ocupação) are done on clean data
        duplicates_before = len(template)
        template = template.drop_duplicates(subset=['COD FORNECEDOR', 'COD DESTINO', 'DESENHO', 'QTDE']).reset_index(drop=True)
    
        # --- Filter out FLECHINHA == 1 AND COD DESTINO == 1080 ---
        rows_before_flechinha_filter = len(template)
        template = template[~((template['FLECHINHA'] == 1) & (template['COD DESTINO'] == '1080'))].reset_index(drop=True)
        rows_removed_flechinha = rows_before_flechinha_filter - len(template)
        if rows_removed_flechinha > 0:
            adicionar_erro(f"{rows_removed_flechinha} linha(s) removida(s) (FLECHINHA=1 e COD DESTINO=1080)", "INFO")
       
        # --- Construção da aba Saturação ---
        df_saturacao = (
            template.groupby(['COD FORNECEDOR', 'FORNECEDOR', 'MDR'], as_index=False)['QTD EMBALAGENS']
            .sum()
            .rename(columns={'MDR': 'EMBALAGEM', 'QTD EMBALAGENS': 'TOTAL DE CXS'})
        )

        # Recupera a coluna VEICULO para cada fornecedor + embalagem
        col_veiculo = template[['COD FORNECEDOR', 'MDR', 'VEICULO']].drop_duplicates()
        col_veiculo = col_veiculo.rename(columns={'MDR': 'EMBALAGEM'})

        df_saturacao = df_saturacao.merge(col_veiculo, on=['COD FORNECEDOR', 'EMBALAGEM'], how='left')

        # Create mappings from db_MDR - filter out NaN values before deduplication
        # to ensure we don't get empty/null values when valid values exist
        db_MDR_valid_paletizavel = db_MDR[db_MDR['CAIXA PLÁSTICA'].notna()]
        mapa_paletizavel = db_MDR_valid_paletizavel.drop_duplicates('MDR').set_index('MDR')['CAIXA PLÁSTICA']
        
        # For CAIXAS POR PALLET, filter out NaN and use the most common value (mode)
        # This prevents picking the first row if it has NaN when other rows have valid values
        db_MDR_valid_cxs = db_MDR[db_MDR['CAIXAS POR PALLET'].notna()]
        
        # Group by MDR and take the mode (most common value) for CAIXAS POR PALLET
        # If multiple modes exist, take the first one
        mapa_cxs_por_pallet = db_MDR_valid_cxs.groupby('MDR')['CAIXAS POR PALLET'].agg(
            lambda x: x.mode()[0] if not x.mode().empty else x.iloc[0]
        )

        df_saturacao['CX_PALETIZÁVEL'] = df_saturacao['EMBALAGEM'].map(mapa_paletizavel).fillna(0).astype(int)
        df_saturacao['CXS_POR_PALLET'] = df_saturacao.apply(
            lambda row: 1 if row['CX_PALETIZÁVEL'] != 1 else (
                mapa_cxs_por_pallet.get(row['EMBALAGEM'], 1) or 1), axis=1
        )
        
        df_saturacao['CXS/PALLETS_TOTAL'] = df_saturacao['TOTAL DE CXS'] / df_saturacao['CXS_POR_PALLET']

        # Mapeia de código do veículo (ex: 4) -> coluna de capacidade no db_MDR (ex: "14 x 2,4 x 2,78")
        mapa_coluna_capacidade = db_veiculos.set_index('COD VEICULO')['VEICULOS'].to_dict()
        
        # --- Vehicle configuration handling ---
        # If usar_manual=True: use global vehicle for all calculations
        # If usar_manual=False: use per-row vehicle from Template (respects FLUXO.xlsx)
        if usar_manual:
            valor_veiculo = db_veiculos.loc[db_veiculos['COD VEICULO'] == veiculo, 'VEICULOS'].iloc[0]
           
        else:
            # When usar_manual=False, we'll use per-row vehicle columns
            # Set valor_veiculo to None to indicate we need per-row lookup
            valor_veiculo = None
            # print(f"\n[INFO] usar_manual=False: Using PER-ROUTE vehicles from Template/FLUXO")
            # unique_vehicles = df_saturacao['VEICULO'].unique()
            # print(f"  * Unique vehicle codes in use: {sorted([int(v) for v in unique_vehicles if pd.notna(v)])}")
            # for v_code in sorted(unique_vehicles):
            #     if pd.notna(v_code):
            #         v_name = mapa_coluna_capacidade.get(int(v_code), "Unknown")
            #         print(f"    - Vehicle {int(v_code)}: {v_name}")
        

        # Garante que os MDRs na base estejam em caixa alta
        db_MDR['MDR'] = db_MDR['MDR'].astype(str).str.upper()

        def obter_veiculo_anterior(cod_veic):
            if cod_veic in [4, 5, 6, 7, 8, 9, 14]:
                return 3
            elif cod_veic in [2, 3, 12, 13, 15, 16, 17, 18]:
                return 1
            elif cod_veic == 1:
                return 10
            elif cod_veic == 10:
                return 11
            elif cod_veic == 11:
                return 11
            return None

        def obter_capacidade_por_linha(row):
            mdr = str(row['EMBALAGEM']).upper()  # Converte para string e caixa alta
            cod_veic = row['VEICULO']
            fornecedor = row['COD FORNECEDOR']
            coluna = mapa_coluna_capacidade.get(cod_veic)

            if not coluna:
                # print(f"[ERRO] Código de veículo {cod_veic} não mapeado.")
                return None
            if coluna not in db_MDR.columns:
                # print(f"[ERRO] Coluna '{coluna}' não encontrada no db_MDR para veículo {cod_veic}")
                return None

            # Try supplier-specific lookup first
            # Handle supplier code equivalence: 800006372 (SAP) = 21544 (IMS)
            supplier_codes = [fornecedor]
            if fornecedor == 800006372:
                supplier_codes.append(21544)
            elif fornecedor == 21544:
                supplier_codes.append(800006372)
            
            # Use exact matching (faster than contains)
            filtro_fornecedor = (db_MDR['MDR'] == mdr) & (db_MDR['CÓD. FORNECEDOR'].isin(supplier_codes))
            capacidade_series_forn = db_MDR.loc[filtro_fornecedor, coluna].dropna()
            
            if not capacidade_series_forn.empty:
                # Found supplier-specific capacity - use HYBRID approach
                capacity_mean = capacidade_series_forn.mean()
                capacity_min = capacidade_series_forn.min()
                capacity_max = capacidade_series_forn.max()
                capacity_mode = capacidade_series_forn.mode().values[0] if not capacidade_series_forn.mode().empty else capacity_mean
                
                # Hybrid: Use MODE if high variance, otherwise MAX
                variance = capacity_max - capacity_min
                # If variance is more than 50% of MODE, use MODE (data likely has outliers)
                if variance > capacity_mode * 0.5:
                    capacity_selected = capacity_mode
                    selection_method = 'MODE'
                else:
                    capacity_selected = capacity_max
                    selection_method = 'MAX'
                
                return capacity_selected
            
            # Fall back to all suppliers for this MDR, use MAX
            filtro = db_MDR['MDR'] == mdr
            capacidade_series = db_MDR.loc[filtro, coluna].dropna()

            if capacidade_series.empty:
                # print(f"[ERRO] Capacidade não encontrada para MDR {mdr} na coluna '{coluna}' (cod veic {cod_veic})")
                return None

            # Use MAX approach to aggregate capacities
            capacity_mean = capacidade_series.mean()
            capacity_min = capacidade_series.min()
            capacity_max = capacidade_series.max()
            capacity_mode = capacidade_series.mode().values[0] if not capacidade_series.mode().empty else capacity_mean
            
            # Use MAX for calculations
            capacity_selected = capacity_max
            selection_method = 'MAX'
            
            return capacity_selected

        def obter_capacidade_por_linha_veic_anterior(row):

            mdr = str(row['EMBALAGEM']).upper()
            cod_veic = int(row['VEICULO'])
            veic_anterior = obter_veiculo_anterior(cod_veic)
           
            if veic_anterior is None :
                # print(f"[INFO] Veículo anterior não definido para código {cod_veic}")
                return None

            coluna = mapa_coluna_capacidade.get(veic_anterior)

            if not coluna:
                # print(f"[ERRO] Código de veículo anterior {veic_anterior} não mapeado.")
                return None
            
            if coluna not in db_MDR.columns:
                # print(f"[ERRO] Coluna '{coluna}' não encontrada no db_MDR para veículo anterior {veic_anterior}")
                return None

            filtro = db_MDR['MDR'].str.contains(mdr)
            capacidade_series = db_MDR.loc[filtro, coluna].dropna()

            if capacidade_series.empty:
                print(
                    f"[ERRO] Capacidade não encontrada para MDR {mdr} na coluna '{coluna}' (veic anterior {veic_anterior})")
                return None

            return capacidade_series.values[0]

        df_saturacao['CAPACIDADE'] = df_saturacao.apply(obter_capacidade_por_linha, axis=1)
        df_saturacao['VEICULO'] = df_saturacao['VEICULO'].fillna(0)
        df_saturacao['VEICULO'] = df_saturacao['VEICULO'].astype(int)         
        df_saturacao['CAPACIDADE_VEIC_ANTERIOR'] = df_saturacao.apply(obter_capacidade_por_linha_veic_anterior, axis=1)

       
        # Converte para numérico, tratando valores não numéricos
        df_saturacao['CAPACIDADE'] = pd.to_numeric(df_saturacao['CAPACIDADE'], errors='coerce')
        df_saturacao['CAPACIDADE_VEIC_ANTERIOR'] = pd.to_numeric(df_saturacao['CAPACIDADE_VEIC_ANTERIOR'], errors='coerce')
        df_saturacao['CXS/PALLETS_TOTAL'] = pd.to_numeric(df_saturacao['CXS/PALLETS_TOTAL'], errors='coerce')
        
        # Calcula saturação apenas onde CAPACIDADE_VEIC_ANTERIOR não é nulo/zero
        df_saturacao['SATURAÇÃO COM VEÍCULO MENOR (%)'] = 0.0
        mask = (df_saturacao['CAPACIDADE_VEIC_ANTERIOR'].notna()) & (df_saturacao['CAPACIDADE_VEIC_ANTERIOR'] > 0)
        df_saturacao.loc[mask, 'SATURAÇÃO COM VEÍCULO MENOR (%)'] = round(
            df_saturacao.loc[mask, 'CXS/PALLETS_TOTAL'] / df_saturacao.loc[mask, 'CAPACIDADE_VEIC_ANTERIOR'] * 100, 2
        )

        bases = set(zip(db_empilhamento['FORNECEDOR'], db_empilhamento['MDR BASE']))
        sobrepostas = set(zip(db_empilhamento['FORNECEDOR'], db_empilhamento['MDR SOBREPOSTA']))
        df_saturacao['EMBALAGEM_BASE'] = df_saturacao.apply(
            lambda row: 1 if (row['FORNECEDOR'], row['EMBALAGEM']) in bases else 0, axis=1)
        df_saturacao['EMBALAGEM_SOBREPOSTA'] = df_saturacao.apply(
            lambda row: 1 if (row['FORNECEDOR'], row['EMBALAGEM']) in sobrepostas else 0, axis=1)

        df_saturacao['CHAVE'] = df_saturacao['COD FORNECEDOR'].astype(str) + '-' + df_saturacao['EMBALAGEM'].astype(str)


        # --- Eficiência de empilhamento por embalagem ---
        # When usar_manual=False, we need to handle multiple vehicles (per-row lookup)
        # When usar_manual=True, we can use a single global vehicle column
        if usar_manual and valor_veiculo is not None:
            # Global vehicle mode: use single efficiency column for all rows
            if valor_veiculo in db_efi.columns:
                mapa_efi = db_efi.drop_duplicates('CHAVE FORNE + MDR').set_index('CHAVE FORNE + MDR')[valor_veiculo]
                df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = df_saturacao['CHAVE'].map(mapa_efi).fillna(1)
            else:
                print(f"[WARNING] Efficiency column '{valor_veiculo}' not found in BD_CADASTRO_MDR_PERDA_COMPRIMENTO")
                df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = 1.0
        else:
            # Per-route vehicle mode: lookup efficiency for each row's vehicle
            def obter_eficiencia_por_linha(row):
                chave = str(row['COD FORNECEDOR']) + '-' + str(row['EMBALAGEM'])
                cod_veic = row['VEICULO']
                coluna_veic = mapa_coluna_capacidade.get(cod_veic)
                
                if not coluna_veic or coluna_veic not in db_efi.columns:
                    return 1.0  # Default efficiency
                
                # Filter db_efi for this CHAVE and get efficiency from appropriate vehicle column
                filtro = db_efi['CHAVE FORNE + MDR'] == chave
                efi_series = db_efi.loc[filtro, coluna_veic].dropna()
                
                if efi_series.empty:
                    return 1.0
                
                return efi_series.values[0]
            
            df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = df_saturacao.apply(obter_eficiencia_por_linha, axis=1)
            df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = df_saturacao['EFICIÊNCIA_COMPRIMENTO'].fillna(1)


        mapa_volume_efi = db_MDR.drop_duplicates('CHAVE EMBALAGENS').set_index('CHAVE EMBALAGENS')['VOLUME']
        df_saturacao['M³ POR EMBALAGEM'] = df_saturacao['CHAVE'].map(mapa_volume_efi) * \
                                            df_saturacao['CXS_POR_PALLET'] * df_saturacao['CXS/PALLETS_TOTAL']

        # --- Cálculo de empilhamento ---
        df_calculo_empilhamento = calcular_empilhamento(df_saturacao, db_empilhamento)

        # --- Saturação final por embalagem ---
        def integrar_saturacao_total(df_sat, df_emp):
            def calcular(row):
                filtro = (df_emp['FORNECEDOR'] == row['COD FORNECEDOR']) & \
                         (df_emp['EMBALAGEM_BASE'] == row['EMBALAGEM'])
                soma_saturacoes = df_emp[filtro]['SATURAÇÃO'].sum()
                # Prevent division by zero
                if pd.isna(row['CAPACIDADE']) or row['CAPACIDADE'] == 0:
                    return 0
                proporcao = row['CXS/PALLETS_TOTAL'] / row['CAPACIDADE']
                result = (proporcao + soma_saturacoes) * row['EFICIÊNCIA_COMPRIMENTO']
                

                
                return result

            df_sat['SATURAÇÃO_TOTAL'] = df_sat.apply(calcular, axis=1)
            # Clean up infinity values
            df_sat['SATURAÇÃO_TOTAL'] = df_sat['SATURAÇÃO_TOTAL'].replace([np.inf, -np.inf], np.nan).fillna(0)
            
            # Prevent division by zero for SATURAÇÃO_POR_MDR
            df_sat['SATURAÇÃO_POR_MDR'] = 0.0
            mask = (df_sat['TOTAL DE CXS'].notna()) & (df_sat['TOTAL DE CXS'] > 0)
            df_sat.loc[mask, 'SATURAÇÃO_POR_MDR'] = df_sat.loc[mask, 'SATURAÇÃO_TOTAL'] / df_sat.loc[mask, 'TOTAL DE CXS']
            df_sat['SATURAÇÃO_POR_MDR'] = df_sat['SATURAÇÃO_POR_MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)
            
            return df_sat
            
            return df_sat

        if not df_calculo_empilhamento.empty:
            df_saturacao = integrar_saturacao_total(df_saturacao, df_calculo_empilhamento)
        else:
            # Prevent division by zero
            df_saturacao['SATURAÇÃO_TOTAL'] = 0.0
            mask = (df_saturacao['CAPACIDADE'].notna()) & (df_saturacao['CAPACIDADE'] > 0)
            df_saturacao.loc[mask, 'SATURAÇÃO_TOTAL'] = df_saturacao.loc[mask, 'CXS/PALLETS_TOTAL'] / df_saturacao.loc[mask, 'CAPACIDADE']
            df_saturacao['SATURAÇÃO_TOTAL'] = df_saturacao['SATURAÇÃO_TOTAL'].replace([np.inf, -np.inf], np.nan).fillna(0)
            
            df_saturacao['SATURAÇÃO_POR_MDR'] = 0.0
            mask = (df_saturacao['TOTAL DE CXS'].notna()) & (df_saturacao['TOTAL DE CXS'] > 0)
            df_saturacao.loc[mask, 'SATURAÇÃO_POR_MDR'] = df_saturacao.loc[mask, 'SATURAÇÃO_TOTAL'] / df_saturacao.loc[mask, 'TOTAL DE CXS']
            df_saturacao['SATURAÇÃO_POR_MDR'] = df_saturacao['SATURAÇÃO_POR_MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)

        # --- Cálculo da SAT por linha ---
        template.loc[:, 'CHAVE'] = template['COD FORNECEDOR'].astype(str) + '-' + template['MDR'].astype(str)
        
        template = template.merge(df_saturacao[['CHAVE', 'SATURAÇÃO_POR_MDR']], on='CHAVE', how='left')
        
        # Clean up SATURAÇÃO_POR_MDR to avoid infinity values
        template['SATURAÇÃO_POR_MDR'] = template['SATURAÇÃO_POR_MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)
        
        template['SAT VOLUME (%)'] = round(template['QTD EMBALAGENS'] * template['SATURAÇÃO_POR_MDR'] * 100, 2)
        
        # Prevent division by zero for SAT PESO calculation
        template['PESO_MAXIMO'] = template['PESO_MAXIMO'].replace(0, np.nan)
        template['SAT PESO (%)'] = round(template['PESO TOTAL'] / template['PESO_MAXIMO'] * 100, 2)
        
        # Clean up any infinity/NaN values in SAT columns
        template['SAT VOLUME (%)'] = template['SAT VOLUME (%)'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['SAT PESO (%)'] = template['SAT PESO (%)'].replace([np.inf, -np.inf], np.nan).fillna(0)
        
        
       
        template.drop(columns=['CHAVE', 'SATURAÇÃO_POR_MDR'], inplace=True)
        df_saturacao.drop(columns=['CHAVE'], inplace=True)

        # --- Criação das variáveis para a tabela final ---
        # Filter to only include rows that should be in calculations
        template_calc = template[template['INCLUDE_IN_CALC']] if 'INCLUDE_IN_CALC' in template.columns else template
        
        ocupacao = template_calc['SAT VOLUME (%)'].sum()
        qtd_veiculos = (ceil(ocupacao / 100))
        volume = template_calc['M³'].sum()
        peso = template_calc['PESO TOTAL'].sum()
        
        # Handle infinity and NaN values in QTD EMBALAGENS before summing
        embalagens_series = template_calc['QTD EMBALAGENS'].replace([np.inf, -np.inf], np.nan).fillna(0)
        embalagens = embalagens_series.sum()

        # Preenche a tree_resumo (que deve ser passada como argumento)
        resumo_dados = [
            ("Ocupação Total", f"{ocupacao:.2f}%"),
            ("Qtd Veículos", qtd_veiculos),
            ("Volume Total", f"{volume:.1f} m³"),
            ("Peso Total", f"{peso:.1f} kg"),
            #("Peso Máximo", f"{peso_maximo:.1f} kg"),
            ("Embalagens", int(embalagens) if np.isfinite(embalagens) else 0),
        ]

        linhas_validas = template[
            (template['DESENHO'].notna()) &
            (template['COD FORNECEDOR'].notna()) &
            (template['QTDE'] > 0)
            ].shape[0]

        linha_qme = template[
            (template['QME'] > 0) &
            (template['QTDE'] > 0)
            ].shape[0]

        # Limpa e atualiza a tabela tree_resumo
        tree_resumo.delete(*tree_resumo.get_children())
        for item in resumo_dados:
            tree_resumo.insert("", END, values=item)


        # --- Atualiza TreeView (Tkinter) with deduplicated data ---
        tree.delete(*tree.get_children())
        tree["columns"] = list(template.columns)
        tree["show"] = "headings"
        
        # Define custom widths for specific columns - optimized for scrolling
        column_widths = {
            'COD FORNECEDOR': 110,
            'FORNECEDOR': 150,
            'COD DESTINO': 90,
            'DESENHO': 90,
            'QTDE': 70,
            'DESCRIÇÃO MATERIAL': 200,
            'MDR': 70,
            'DESCRIÇÃO DA EMBALAGEM': 160,
            'QME': 60,
            'QTD EMBALAGENS': 110,
            'TIPO SATURACAO': 90,
            'VEICULO': 70,
            'MOT': 60,
            'FLECHINHA': 80,
            'M³': 60,
            'PESO MAT': 90,
            'PESO MDR': 90,
            'PESO TOTAL': 100,
            'PESO_MAXIMO': 110,
            'SAT VOLUME (%)': 110,
            'SAT PESO (%)': 100,
            'COD IMS': 90,
            'INCLUDE_IN_CALC': 110
        }
        
        for col in template.columns:
            tree.heading(col, text=col)
            width = column_widths.get(col, 90)  # Default 90 if not specified
            tree.column(col, width=width, anchor="center", stretch=False)
        for _, row in template.iterrows():
            tree.insert("", END, values=list(row))

        desenhar_caminhoes(canvas_caminhoes, ocupacao, caminhao_img)
        
        # --- Second deduplication check before Excel write ---
        # Safety check in case any operations after calculations reintroduced duplicates
        duplicates_before_write = len(template)
        template = template.drop_duplicates(subset=['COD FORNECEDOR', 'COD DESTINO', 'DESENHO', 'QTDE']).reset_index(drop=True)
        duplicates_removed_write = duplicates_before_write - len(template)
        if duplicates_removed_write > 0:
            adicionar_erro(f"{duplicates_removed_write} linha(s) duplicada(s) removida(s) antes de salvar VIAJANTE", "INFO")
            print(f"[INFO] Removed {duplicates_removed_write} additional duplicates before Excel write")
        
        # --- Filter based on COD IMS and COD DESTINO ---
        # COD IMS can contain multiple values separated by '/' (e.g., "800002365/800036730")
        # Remove rows where: COD FORNECEDOR is contained in COD IMS AND COD DESTINO != 1046
        # Keep rows where: COD FORNECEDOR is NOT in COD IMS, OR (is in COD IMS AND COD DESTINO = 1046)
        if 'COD IMS' in template.columns:
            rows_before_filter = len(template)
            
            # Normalize COD IMS to string for comparison (already normalized COD FORNECEDOR and COD DESTINO above)
            template['COD IMS'] = template['COD IMS'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
            # Convert COD FORNECEDOR to string for comparison
            template['COD FORNECEDOR_STR'] = template['COD FORNECEDOR'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
            # Function to check if COD FORNECEDOR is in COD IMS (handling / separator)
            def fornecedor_in_ims(row):
                cod_forn = str(row['COD FORNECEDOR_STR'])
                cod_ims = str(row['COD IMS'])
                if pd.isna(cod_ims) or cod_ims == '' or cod_ims == 'nan' or cod_ims == '0':
                    return False
                # Split COD IMS by '/' and check if COD FORNECEDOR is in the list
                ims_values = [v.strip() for v in cod_ims.split('/')]
                return cod_forn in ims_values
            
            # Build filter mask: keep if (NOT in IMS) OR (in IMS AND destino = 1046)
            template['IS_IN_IMS'] = template.apply(fornecedor_in_ims, axis=1)
            mask_keep = (~template['IS_IN_IMS']) | ((template['IS_IN_IMS']) & (template['COD DESTINO'] == '1046'))
            
            # Apply filter
            template = template[mask_keep].copy()
            
            # Clean up temporary columns
            template = template.drop(columns=['COD FORNECEDOR_STR', 'IS_IN_IMS'])
            
            rows_removed = rows_before_filter - len(template)
            if rows_removed > 0:
                adicionar_erro(f"{rows_removed} linha(s) removida(s) (COD FORNECEDOR in COD IMS e COD DESTINO != 1046)", "INFO")
        
        # --- Exporta para Excel formatado ---
        with pd.ExcelWriter('VIAJANTE.xlsx', engine='openpyxl') as writer:
            template.to_excel(writer, sheet_name='Template Completo', index=False)
            df_saturacao.to_excel(writer, sheet_name='Saturação', index=False)
            df_calculo_empilhamento.to_excel(writer, sheet_name='Calculo Empilhamento', index=False)

            header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            header_align = Alignment(horizontal='center', vertical='center')

            for sheet_name in ['Template Completo', 'Saturação', 'Calculo Empilhamento']:
                ws = writer.sheets[sheet_name]
                for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                    largura = max(len(str(cell.value) or '') for cell in col) + 2
                    ws.column_dimensions[get_column_letter(col_num)].width = largura
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align

            if 'MDR' in template.columns:
                pn_nao_cadastrados = template[
                    template['MDR'].isna() | (template['MDR'].astype(str).str.strip() == '')
                ].copy()

                # select only the requested columns if they exist in the dataframe
                cols_to_keep = ['COD FORNECEDOR', 'FORNECEDOR', 'COD DESTINO', 'DESENHO']
                existing_cols = [c for c in cols_to_keep if c in pn_nao_cadastrados.columns]

                if not pn_nao_cadastrados.empty and existing_cols:
                    pn_nao_cadastrados = pn_nao_cadastrados[existing_cols]
                    pn_nao_cadastrados.drop_duplicates(subset=["DESENHO"], inplace=True)
                    pn_nao_cadastrados.to_excel(writer, sheet_name='PN Não Cadastrados', index=False)
                    
                    # Log de PNs não cadastrados
                    qtd_pn_faltando = len(pn_nao_cadastrados)
                    adicionar_erro(f"{qtd_pn_faltando} desenho(s) sem MDR cadastrado. Verifique a aba 'PN Não Cadastrados'. do viajante", "AVISO")

    except Exception as e:
        adicionar_erro(f"Erro crítico ao processar informações: {str(e)}", "ERRO")
        print(f"Erro: {e}")
        traceback.print_exc()


def consolidar_dados(use_manual=False, manual_veiculo=None):
    # === CONFIGURATION: Static Cargas Adjustment ===
    use_static_adjustment = 1  # Set to 1 to activate, 0 to deactivate
    
    # Static adjustment values per supplier (applied after normal cargas calculation)
    static_adjustments = {
        '800002365': +1,
        '800030834': -1,
        # '800048577': -1,
        # '800006356': +1,
        '800046699': -1,
        '800014209': -1,
        '800045273': -1,
    }
    # ================================================
    
    # Carrega os dados
    fluxos_path = os.path.join(caminho_base, "BD", "FLUXO.xlsx")
    fluxos = pd.read_excel(fluxos_path, sheet_name='FLUXOS')
    
    # Ensure COD IMS column exists in fluxos
    if 'COD IMS' not in fluxos.columns:
        fluxos['COD IMS'] = ""
    
    template = pd.read_excel('VIAJANTE.xlsx', sheet_name='Template Completo')

    # Filtra linhas com quantidade válida e prepara as colunas
    template = template[template['QTDE'] > 0].copy()
    
    # Ensure COD IMS column exists (for backward compatibility)
    if 'COD IMS' not in template.columns:
        template['COD IMS'] = ""
    
    # Clean up COD FORNECEDOR - remove .0 suffix before converting to string
    template['COD FORNECEDOR'] = template['COD FORNECEDOR'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    
    template['COD DESTINO'] = template['COD DESTINO'].astype(str).str.strip()
    template['COD DESTINO'] = template['COD DESTINO'].str.replace(r'\.0$', '', regex=True)
    
    # Normalize FLUXO data types - clean up .0 suffix
    fluxos['COD FORNECEDOR'] = fluxos['COD FORNECEDOR'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    fluxos['COD DESTINO'] = fluxos['COD DESTINO'].astype(str).str.strip()
    fluxos['COD DESTINO'] = fluxos['COD DESTINO'].str.replace(r'\.0$', '', regex=True)
    
    
    # Handle COD IMS if present - clean up .0 suffix
    if 'COD IMS' in template.columns:
        template['COD IMS'] = template['COD IMS'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    if 'COD IMS' in fluxos.columns:
        fluxos['COD IMS'] = fluxos['COD IMS'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

    # Clean up FORNECEDOR - remove .0 suffix and handle NaN
    template['FORNECEDOR'] = template['FORNECEDOR'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)

    # --- Filter based on COD IMS and COD DESTINO ---
    # COD IMS can contain multiple values separated by '/' (e.g., "800002365/800036730")
    # Remove rows where: COD FORNECEDOR is contained in COD IMS AND COD DESTINO != 1046
    # Keep rows where: COD FORNECEDOR is NOT in COD IMS, OR (is in COD IMS AND COD DESTINO = 1046)
    if 'COD IMS' in template.columns:
        rows_before_filter = len(template)
        
        # Function to check if COD FORNECEDOR is in COD IMS (handling / separator)
        def fornecedor_in_ims(row):
            cod_forn = str(row['COD FORNECEDOR'])
            cod_ims = str(row['COD IMS'])
            if pd.isna(cod_ims) or cod_ims == '' or cod_ims == 'nan' or cod_ims == '0':
                return False
            # Split COD IMS by '/' and check if COD FORNECEDOR is in the list
            ims_values = [v.strip() for v in cod_ims.split('/')]
            return cod_forn in ims_values
        
        # Build filter mask: keep if (NOT in IMS) OR (in IMS AND destino = 1046)
        template['IS_IN_IMS'] = template.apply(fornecedor_in_ims, axis=1)
        mask_keep = (~template['IS_IN_IMS']) | ((template['IS_IN_IMS']) & (template['COD DESTINO'] == '1046'))
        
        # Apply filter
        template = template[mask_keep].copy()
        
        # Clean up temporary column
        template = template.drop(columns=['IS_IN_IMS'])
        
        rows_removed = rows_before_filter - len(template)
        if rows_removed > 0:
            print(f"[INFO] consolidar_dados: Removed {rows_removed} rows (COD FORNECEDOR in COD IMS and COD DESTINO != 1046)")

    # If user forced a manual vehicle, override the template VEICULO column
    if use_manual and manual_veiculo is not None:
        try:
            template['VEICULO'] = int(manual_veiculo)
        except Exception:
            template['VEICULO'] = manual_veiculo
        # Try to persist the overridden Template Completo back into VIAJANTE.xlsx
        try:
            via_path = os.path.join(caminho_base, 'VIAJANTE.xlsx')
            if os.path.exists(via_path):
                out_template = template.copy()
                try:
                    # map codes to names using code_to_vehicle_name if available
                    def map_name2(v):
                        try:
                            return code_to_vehicle_name.get(int(v), v)
                        except Exception:
                            return code_to_vehicle_name.get(v, v)
                    out_template['VEICULO'] = out_template['VEICULO'].apply(map_name2)
                except Exception:
                    pass
                with pd.ExcelWriter(via_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    out_template.to_excel(writer, sheet_name='Template Completo', index=False)
        except Exception:
            # Non-fatal: continue even if we can't write back to the file
            pass

    # Build vehicle code -> display name mapping (if BD/VEÍCULOS exists)
    code_to_vehicle_name = {}
    try:
        veic_path = os.path.join(caminho_base, 'BD', 'VEÍCULOS.xlsx')
        if os.path.exists(veic_path):
            db_veic = pd.read_excel(veic_path)
            cols = {c.strip().upper(): c for c in db_veic.columns}
            code_col = None
            name_col = None
            # Prefer explicit 'DESCRICAO' column for vehicle display name
            for up, orig in cols.items():
                if up == 'DESCRICAO':
                    name_col = orig
                if 'COD' in up and 'VEIC' in up:
                    code_col = orig
            # Fallbacks
            if code_col is None:
                for up, orig in cols.items():
                    if 'COD' in up:
                        code_col = orig
                        break
            if name_col is None:
                # try any column containing 'VEIC' or 'VEICULO'
                for up, orig in cols.items():
                    if 'VEIC' in up or 'VEICULO' in up:
                        name_col = orig
                        break
            if name_col is None and len(db_veic.columns) > 1:
                name_col = db_veic.columns[1]

            if code_col and name_col:
                for _, r in db_veic.iterrows():
                    raw_code = r.get(code_col)
                    try:
                        key = int(raw_code)
                    except Exception:
                        try:
                            key = int(float(str(raw_code)))
                        except Exception:
                            key = str(raw_code).strip()
                    code_to_vehicle_name[key] = str(r.get(name_col, '')).strip()
    except Exception:
        code_to_vehicle_name = {}

    def normalizar_codigos(campo):
        if pd.isna(campo):
            return []
        return [c.strip() for c in re.split(r'\s*[,/]\s*', str(campo).strip()) if c.strip()]

    dados_volume = []

    processed_suppliers = set()
    all_template_suppliers = set()

    # Collect all unique destination codes from template, cleaning them.
    # It filters out 'nan' strings, as '.0' normalization now happens earlier.
    s = template['COD DESTINO'].dropna()
    s = s[s.str.lower() != 'nan']
    all_cod_destinos = set(s.unique())
    
    
    for cod_dest in all_cod_destinos:
        # Find all rows in template that match this exact destination code
        mask_template = template['COD DESTINO'] == cod_dest
        subset_template = template[mask_template]

        # Build set of supplier codes - use COD IMS first, then COD FORNECEDOR
        fornecedores_template_set = set()
        for _, row in subset_template.iterrows():
            # If COD IMS exists and is not empty/nan, use it; otherwise use COD FORNECEDOR
            if 'COD IMS' in template.columns and pd.notna(row.get('COD IMS')) and str(row.get('COD IMS')).strip() not in ['', 'nan', 'None']:
                fornecedores_template_set.update(normalizar_codigos(row['COD IMS']))
            else:
                fornecedores_template_set.update(normalizar_codigos(row['COD FORNECEDOR']))

    

        all_template_suppliers.update(fornecedores_template_set)

        # Find routes that match this exact destination code
        mask_fluxo = fluxos['COD DESTINO'] == cod_dest
        rotas_destino = fluxos[mask_fluxo]
     
        
        for _, rota in rotas_destino.iterrows():
            cod_fluxo = rota['COD FLUXO']
            destino = rota['NOME DESTINO']
            veiculo = rota['VEICULO PRINCIPAL']
            tipo_saturacao = rota['TIPO SATURACAO']
            transportadora = rota['TRANSPORTADORA']
            
            # Get supplier codes from route - check both COD IMS and COD FORNECEDOR
            fornecedores_rota = set()
            if 'COD IMS' in fluxos.columns and pd.notna(rota.get('COD IMS')):
                fornecedores_rota.update(normalizar_codigos(rota['COD IMS']))
            fornecedores_rota.update(normalizar_codigos(rota['COD FORNECEDOR']))

            # Find common suppliers between template and route
            fornecedores_comuns = [f for f in fornecedores_rota if f in fornecedores_template_set]
             
            if fornecedores_comuns:
                processed_suppliers.update(fornecedores_comuns)
                # Find template rows that match - check both COD IMS and COD FORNECEDOR
    
                def row_matches_suppliers(row):
                    row_codes = set()
                    if 'COD IMS' in template.columns and pd.notna(row.get('COD IMS')) and str(row.get('COD IMS')).strip() not in ['', 'nan', 'None']:
                        row_codes.update(normalizar_codigos(row['COD IMS']))
                    
                    row_codes.update(normalizar_codigos(row['COD FORNECEDOR']))
                    return any(f in row_codes for f in fornecedores_comuns)
                
                
                
                mask_fornecedor = subset_template.apply(row_matches_suppliers, axis=1)
                
                # Filter by INCLUDE_IN_CALC to exclude invalid CT PNs
                if 'INCLUDE_IN_CALC' in subset_template.columns:
                    mask_fornecedor = mask_fornecedor & subset_template['INCLUDE_IN_CALC']
                
                linhas_rota = subset_template[mask_fornecedor]


                volume_total = linhas_rota['M³'].sum()
                peso_total = linhas_rota['PESO TOTAL'].sum()
                embalagens_total = linhas_rota['QTD EMBALAGENS'].sum()

                if tipo_saturacao.upper() == 'VOLUME':
                    saturacao_total = linhas_rota['SAT VOLUME (%)'].sum()
                else:
                    saturacao_total = linhas_rota['SAT PESO (%)'].sum()

                nomes_fornecedores = linhas_rota[['COD FORNECEDOR', 'FORNECEDOR']].drop_duplicates()
                nomes_fornecedores['COD FORNECEDOR'] = nomes_fornecedores['COD FORNECEDOR'].astype(str)
                
                
                # Get supplier names in order
                nomes_ordenados = []
                for f in fornecedores_comuns:
                    matching = nomes_fornecedores[nomes_fornecedores['COD FORNECEDOR'].apply(lambda x: f in normalizar_codigos(x))]
                                            
                    if not matching.empty:
                        nomes_ordenados.append(matching.iloc[0]['FORNECEDOR'])
                        
                        
                cargas = math.ceil(saturacao_total / 100) if saturacao_total > 0 else 0

                # Apply static adjustment if enabled
                if use_static_adjustment == 1:
                    for supplier_code in fornecedores_comuns:
                        if supplier_code in static_adjustments:
                            adjustment = static_adjustments[supplier_code]
                            cargas_original = cargas
                            cargas = max(0, int(cargas + adjustment))  # Ensure cargas doesn't go negative and is integer
                            # if cargas != cargas_original:
                            #     print(f"[STATIC ADJUSTMENT] Supplier {supplier_code}: {cargas_original} → {cargas} (adjustment: {adjustment:+g})")

                # Coluna de Sugestão
                saturacao_residual = math.ceil(saturacao_total % 100)
                if cargas > 0 and saturacao_residual <= 20:
                    sugestao = "Cortar coleta do último veículo"
                elif cargas > 0 and saturacao_residual <= 50:
                    sugestao = "Alterar último veículo para menor porte"
                else:
                    sugestao = "Manter coleta"

                

                # Apuração de MDR
                coluna_sat = 'SAT VOLUME (%)' if tipo_saturacao.upper() == 'VOLUME' else 'SAT PESO (%)'
                
                total_desenhos = linhas_rota['DESENHO'].nunique()
                desenhos_apurados = linhas_rota[linhas_rota[coluna_sat].fillna(0) > 0]['DESENHO'].nunique()

                perc_mdr = round((desenhos_apurados / total_desenhos) * 100, 1) if total_desenhos else 0.0
                
                if perc_mdr != 0:
                    # If user elected to force a manual vehicle for the whole run,
                    # override the route vehicle with the provided manual code.
                    veiculo_final = manual_veiculo if use_manual and manual_veiculo is not None else veiculo

                    # Convert code to display name when possible so the Excel shows names
                    veiculo_display = veiculo_final
                    try:
                        # try numeric code
                        code_int = int(veiculo_final)
                        veiculo_display = code_to_vehicle_name.get(code_int, veiculo_final)
                    except Exception:
                        # leave as-is (could already be a name)
                        veiculo_display = veiculo_final
                        
                        
                    dados_volume.append({
                        'COD FLUXO': cod_fluxo,
                        'COD DESTINO': cod_dest,
                        'DESTINO': destino,
                        'CÓDIGOS FORNECEDORES': '/'.join(fornecedores_comuns),
                        'FORNECEDORES NA ROTA': ', '.join(nomes_ordenados),
                        'VEÍCULO': veiculo_display,
                        'TECNOLOGIA': rota['TECNOLOGIA'],
                        'MOT': rota['MOT'],
                        'TRANSPORTADORA': transportadora,
                        'TIPO DE SATURAÇÃO': tipo_saturacao,
                        'VOLUME TOTAL (m³)': round(volume_total, 3),
                        'PESO TOTAL (kg)': round(peso_total, 1),
                        'EMBALAGENS TOTAL': int(embalagens_total),
                        'SATURAÇÃO TOTAL (%)': round(saturacao_total, 2),
                        'CARGAS': cargas,
                        'SUGESTÃO': sugestao,
                        '% MDRs APURADOS': perc_mdr
                    })
     
    missing = all_template_suppliers - processed_suppliers
    # print(f"Suppliers in template but not processed: {missing}")
    
    df_volume = pd.DataFrame(dados_volume)
    # Do NOT drop_duplicates - CT and FTL routes for same supplier are separate!
    
   
   
    
    df_volume.to_excel('Volume_por_rota.xlsx', index=False)
    
#tree = ttk.Treeview()
#tree_resumo = ttk.Treeview()
#completar_informacoes(tree,3, tree_resumo)


